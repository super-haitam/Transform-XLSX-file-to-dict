import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import pprint

# Color where there is a free day
greenColor = "FF92d050"

# Import Data
fileName = input("Type the file path: ")
df = pd.read_excel(fileName)
book = load_workbook(fileName)
sheet = book.active

# Confirm 'df' has the same col structure as 'book'
for i in range(len(df.columns)):
    if sheet.cell(1, i+1).value != df.columns[i]:
        raise ValueError(f"The column name number {i+1} in XLSX file does not match the pandas file.\n\t{sheet.cell(1, i+1).value} != {pd.columns[i]}")

neededAttributes = ["Day", "Time Slot", "Lesson", "Location"]

# Ask the user for what the heads mean
_ = "".join([f"\n\t{i+1}. {nAtt}" for i, nAtt in enumerate(neededAttributes)])
print(f"We only need 4 attributes:{_}")
print(f"The available attributes are {list(df.columns)}")
print("Write what these correspond to in the given XLSX file (in order only comma in between): ")
attributes = input().split(",")
# assert len(attributes) == len(df.columns)

# Check if all attributes are present
for att in attributes:
    if att not in df.columns:
        raise ValueError(f"We think you misspelled '{att}', please try again.")

# Attribute - column dict
attColDict = {neededAttributes[i] : attributes[i] for i in range(len(attributes))}

# Transform the XLSX file to dictionary as follows:
    # {
    #     "(Day 1)": [
    #         {"startTime": "12H00", "endTime": "13H30", "Lesson": "Physics", "Location": "Amphi ASNI"},
    #         {"startTime": "13H30", "endTime": "15H00", "Lesson": "Physics", "Location": "Amphi ASNI"}
    #     ],
    #     "(Day 2)": ...
    # }
# The time format will be in datetime format
dictionary = {}
rowCounter = 2
lastDay = ""
for _ in range(df.shape[0]):
    # Update lastDay as soon as a new day is found
    currentDay = df[attColDict["Day"]][rowCounter - 2]
    print(currentDay)
    if currentDay is not np.nan:
        lastDay = currentDay

    # Check if the time slot has green color or if it is empty, skip it
    currentTimeSlotCell = sheet.cell(rowCounter, list(df.columns).index( attColDict["Time Slot"])+1)
    if currentTimeSlotCell.fill.fgColor.rgb == greenColor or not currentTimeSlotCell.value:
        print("Green or empty")
        rowCounter += 1
        continue
    
    timeSlotLst =   df[attColDict["Time Slot"]][rowCounter - 2].split(" à ")
    lesson =        df[attColDict["Lesson"]][rowCounter - 2]
    location =      df[attColDict["Location"]][rowCounter -2]

    if lastDay not in dictionary:
        dictionary[lastDay] = []

    dictionary[lastDay].append(
        {
            "startTime": timeSlotLst[0],
            "endTime": timeSlotLst[1],
            "lesson": lesson,
            "location": location
        }
    )

    rowCounter += 1

