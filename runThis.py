from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import json


class XLSXtoDict:
    # Color where there is a free day
    greenColor = "FF92d050"
    neededAttributes = ["Day", "Time Slot", "Lesson", "Location", "Number", "Notice"]

    def __init__(self, filename: str, save: bool):
        self.filename = filename
        self.save = save
        self.importData()
        self.associateColumns()

    def importData(self):
        # Import Data
        self.df = pd.read_excel(self.filename)
        book = load_workbook(self.filename)
        self.sheet = book.active

        # Confirm 'df' has the same col structure as 'book'
        for i in range(len(self.df.columns)):
            if self.sheet.cell(1, i+1).value != self.df.columns[i]:
                raise ValueError(f"The column name number {i+1} in XLSX file does not match the pandas file.\n\t{self.sheet.cell(1, i+1).value} != {self.df.columns[i]}")


    def associateColumns(self) -> dict:
        # Ask the user for what the heads mean
        _ = "".join([f"\n\t{i+1}. {nAtt}" for i, nAtt in enumerate(self.neededAttributes)])
        print(f"We only need 6 attributes:{_}")
        print(f"The available attributes are {list(self.df.columns)}")
        print("Write what these correspond to in the given XLSX file (in order only comma in between): ")
        attributes = input().split(",")
        # assert len(attributes) == len(self.df.columns)

        # Check if all attributes are present
        for att in attributes:
            if att not in self.df.columns:
                raise ValueError(f"We think you misspelled '{att}' as it is not one of the XLSX file columns, please try again.")

        # Attribute - column dict
        self.attColDict = {self.neededAttributes[i] : attributes[i] for i in range(len(attributes))}
    

    def saveToJson(self, data: dict):
        with open("./output/save.json", "w") as f:
            json.dump(data, f, indent=4)

        print("File saved successfully!")


    def transform(self) -> dict:
        # Transform the XLSX file to dictionary as follows:
            # {
            #     "(Day 1)": [
            #         {"startTime": "12H00", "endTime": "13H30", "Lesson": "Physics", "Location": "Amphi ASNI", ...},
            #         {"startTime": "13H30", "endTime": "15H00", "Lesson": "Physics", "Location": "Amphi ASNI", ...}
            #     ],
            #     "(Day 2)": ...
            # }
        # The time format will be in datetime format
        lastTimeSlotIsGreen = False
        dictionary = {}
        rowCounter = 2
        lastDay = ""
        lastTimeSlot = ""
        for _ in range(self.df.shape[0]):
            # Update lastDay as soon as a new day is found
            currentDay = self.df[self.attColDict["Day"]][rowCounter - 2]
            if currentDay is not np.nan:
                lastDay = currentDay

            # Check if the time slot has green color, skip it
            currentTimeSlotCell = self.sheet.cell(rowCounter, list(self.df.columns).index( self.attColDict["Time Slot"])+1)
            if currentTimeSlotCell.fill.fgColor.rgb == self.greenColor:
                lastTimeSlotIsGreen = True
                rowCounter += 1
                continue
            
            # print(lastDay)
            # print(currentTimeSlotCell.coordinate, currentTimeSlotCell.value, bool(currentTimeSlotCell.value))
            # Check if the time slot is empty and the last time slot is green
            if not currentTimeSlotCell.value and lastTimeSlotIsGreen:
                lastTimeSlotIsGreen = False
                rowCounter += 1
                continue

            if currentTimeSlotCell.value:
                print(f"Assigning {currentTimeSlotCell.value} to lastTimeSlot")
                lastTimeSlot = currentTimeSlotCell.value
                lastTimeSlotCell = currentTimeSlotCell

            print(lastTimeSlot)
            timeSlotLst =   lastTimeSlot.split(" à ")
            lesson =        self.df[self.attColDict["Lesson"]][rowCounter - 2]
            location =      self.df[self.attColDict["Location"]][rowCounter -2]
            number =        self.df[self.attColDict["Number"]][rowCounter -2]
            notice =        self.df[self.attColDict["Notice"]][rowCounter -2]

            if lastDay not in dictionary:
                dictionary[lastDay] = []

            dictionary[lastDay].append(
                {
                    "startTime": timeSlotLst[0], 
                    "endTime": timeSlotLst[1],
                    "lesson": lesson,
                    "location": location if location is not np.nan else "(No Location Info Provided)",
                    "number": number if number is not np.nan else "(No Number Provided)",
                    "notice": notice if (notice is not np.nan and notice != "\u00a0") else "(No Notice Provided)"
                }
            )

            rowCounter += 1
        
        if self.save:
            self.saveToJson(dictionary)

        return dictionary


if __name__ == "__main__":
    fileName = input("Type the file path: ")
    instance = XLSXtoDict(fileName, True)
    instance.transform()